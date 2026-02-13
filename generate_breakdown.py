#!/usr/bin/env python3
"""
GROZA Development Stages Breakdown Generator
Generates an Excel file with POC, MVP, and Full Product scope and cost calculation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_groza_breakdown():
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # =====================
    # Summary Sheet
    # =====================
    ws_summary = wb.create_sheet("Summary")
    
    # Header styling
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    sub_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    sub_header_font = Font(color="FFFFFF", bold=True, size=11)
    stage_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    stage_font = Font(color="FFFFFF", bold=True, size=11)
    total_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    total_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws_summary['A1'] = 'GROZA Development Breakdown & Cost Calculator'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:F1')
    
    # Hourly rate input
    ws_summary['A3'] = 'Hourly Rate (€):'
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['B3'] = 100  # Default value
    ws_summary['B3'].number_format = '€#,##0.00'
    ws_summary['B3'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    ws_summary['B3'].font = Font(bold=True, size=12)
    
    # Headers
    row = 5
    headers = ['Stage', 'Category', 'Hours', 'Cost']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data structure
    stages = {
        'POC (Proof of Concept)': {
            'Wireframes & Design': 40,
            'Authentication & Authorization': 60,
            'Company Management (Basic)': 80,
            'CASE Management (Basic)': 60,
            'QR Code Generation': 40,
            'QR Validation (Basic)': 40,
            'Basic Dashboard': 40,
            'Testing & Bug Fixes': 40,
        },
        'MVP (Minimum Viable Product)': {
            'Enhanced UI/UX': 80,
            'Company Onboarding Flow': 60,
            'Employee & Vehicle Management': 80,
            'Document Upload & Storage': 60,
            'Permit Generation (Auto)': 100,
            'DSP Manual Approval': 80,
            'Enhanced QR Validation': 60,
            'Notifications System': 60,
            'Audit Logs': 40,
            'Testing & QA': 80,
        },
        'Full Product': {
            'Inspector Dashboard & Reports': 120,
            'Advanced Analytics': 100,
            'Change Request Workflow': 80,
            'Multi-tenant Architecture': 120,
            'Document OCR & Auto-extraction': 160,
            'Mobile App (iOS/Android)': 400,
            'Advanced Security & Compliance': 80,
            'Performance Optimization': 80,
            'Comprehensive Testing': 120,
            'Documentation & Training': 80,
            'Deployment & DevOps': 80,
        }
    }
    
    row += 1
    start_row = row
    
    for stage_name, categories in stages.items():
        # Stage header
        cell = ws_summary.cell(row=row, column=1)
        cell.value = stage_name
        cell.font = stage_font
        cell.fill = stage_fill
        ws_summary.merge_cells(f'A{row}:B{row}')
        
        # Stage total hours
        stage_hours_cell = ws_summary.cell(row=row, column=3)
        stage_hours = sum(categories.values())
        stage_hours_cell.value = stage_hours
        stage_hours_cell.font = stage_font
        stage_hours_cell.fill = stage_fill
        stage_hours_cell.alignment = Alignment(horizontal='center')
        
        # Stage total cost
        stage_cost_cell = ws_summary.cell(row=row, column=4)
        stage_cost_cell.value = f'=B3*C{row}'
        stage_cost_cell.font = stage_font
        stage_cost_cell.fill = stage_fill
        stage_cost_cell.number_format = '€#,##0.00'
        stage_cost_cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Categories
        for category, hours in categories.items():
            ws_summary.cell(row=row, column=1).value = ''
            ws_summary.cell(row=row, column=2).value = category
            ws_summary.cell(row=row, column=3).value = hours
            ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            cost_cell = ws_summary.cell(row=row, column=4)
            cost_cell.value = f'=B3*C{row}'
            cost_cell.number_format = '€#,##0.00'
            cost_cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        row += 1  # Empty row between stages
    
    # Grand total
    ws_summary.cell(row=row, column=1).value = 'TOTAL'
    ws_summary.cell(row=row, column=1).font = total_font
    ws_summary.cell(row=row, column=1).fill = total_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    
    total_hours_cell = ws_summary.cell(row=row, column=3)
    total_hours_cell.value = f'=SUM(C{start_row}:C{row-1})'
    total_hours_cell.font = total_font
    total_hours_cell.fill = total_fill
    total_hours_cell.alignment = Alignment(horizontal='center')
    
    total_cost_cell = ws_summary.cell(row=row, column=4)
    total_cost_cell.value = f'=B3*C{row}'
    total_cost_cell.font = total_font
    total_cost_cell.fill = total_fill
    total_cost_cell.number_format = '€#,##0.00'
    total_cost_cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 20
    
    # =====================
    # Detailed Breakdown Sheet
    # =====================
    ws_detail = wb.create_sheet("Detailed Breakdown")
    
    # Title
    ws_detail['A1'] = 'GROZA - Detailed Development Stages'
    ws_detail['A1'].font = Font(bold=True, size=14)
    ws_detail.merge_cells('A1:D1')
    
    row = 3
    
    # POC Section
    ws_detail.cell(row=row, column=1).value = 'POC (Proof of Concept)'
    ws_detail.cell(row=row, column=1).font = Font(bold=True, size=13, color="3498DB")
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Goal:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    ws_detail.cell(row=row, column=2).value = 'Validate core concept with clickable wireframes and basic functionality'
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Scope:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    poc_scope = [
        '• Basic authentication (login/logout)',
        '• Simple company registration',
        '• Create CASE with minimal data',
        '• Generate QR code for CASE',
        '• Scan QR and validate company status',
        '• Role-based dashboards (static)',
        '• No document upload',
        '• No approval workflows',
        '• Mock data for testing'
    ]
    
    for item in poc_scope:
        ws_detail.cell(row=row, column=2).value = item
        row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Deliverables:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Clickable wireframes (Figma/Penpot)'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Working prototype with core flows'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Basic database schema'
    row += 2
    
    # MVP Section
    ws_detail.cell(row=row, column=1).value = 'MVP (Minimum Viable Product)'
    ws_detail.cell(row=row, column=1).font = Font(bold=True, size=13, color="27AE60")
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Goal:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    ws_detail.cell(row=row, column=2).value = 'Production-ready core features for initial rollout'
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Scope:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    mvp_scope = [
        '• Complete authentication & user management',
        '• Company onboarding with approval workflow',
        '• Employee & vehicle management',
        '• Document upload & validation',
        '• Automatic permit generation (transport, mortuary passport)',
        '• DSP manual approval for embalming permits',
        '• Enhanced QR validation with all checks',
        '• Email notifications',
        '• Audit logs for all actions',
        '• Company status management (GREEN/RED)',
        '• Basic reporting'
    ]
    
    for item in mvp_scope:
        ws_detail.cell(row=row, column=2).value = item
        row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Deliverables:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Production-ready web application'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• All core workflows functional'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• User documentation'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Deployed on staging/production'
    row += 2
    
    # Full Product Section
    ws_detail.cell(row=row, column=1).value = 'Full Product'
    ws_detail.cell(row=row, column=1).font = Font(bold=True, size=13, color="E74C3C")
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Goal:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    ws_detail.cell(row=row, column=2).value = 'Complete platform with advanced features and mobile support'
    row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Scope:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    full_scope = [
        '• Inspector dashboard with advanced analytics',
        '• Exportable reports (PDF, Excel)',
        '• Company change request workflow',
        '• Multi-tenant architecture',
        '• OCR for document auto-extraction',
        '• Mobile apps (iOS & Android) for validators',
        '• Advanced security (2FA, audit logs)',
        '• Performance optimization for scale',
        '• Comprehensive testing & QA',
        '• Complete documentation',
        '• Training materials',
        '• CI/CD pipeline'
    ]
    
    for item in full_scope:
        ws_detail.cell(row=row, column=2).value = item
        row += 1
    
    ws_detail.cell(row=row, column=1).value = 'Deliverables:'
    ws_detail.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Complete web platform'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• iOS & Android mobile apps'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Admin & user documentation'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Training materials & videos'
    row += 1
    ws_detail.cell(row=row, column=2).value = '• Production deployment'
    
    ws_detail.column_dimensions['A'].width = 20
    ws_detail.column_dimensions['B'].width = 80
    
    # =====================
    # Comparison Sheet
    # =====================
    ws_compare = wb.create_sheet("Stage Comparison")
    
    ws_compare['A1'] = 'Development Stage Comparison'
    ws_compare['A1'].font = Font(bold=True, size=14)
    ws_compare.merge_cells('A1:D1')
    
    # Headers
    row = 3
    comparison_headers = ['Aspect', 'POC', 'MVP', 'Full Product']
    for col, header in enumerate(comparison_headers, 1):
        cell = ws_compare.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    
    comparisons = [
        ('Timeline', '4-6 weeks', '3-4 months', '8-12 months'),
        ('Team Size', '1-2 developers', '3-4 developers', '5-8 developers'),
        ('Users', 'Internal testing only', '50-100 users', 'Unlimited'),
        ('Features', '20% core features', '80% core features', '100% all features'),
        ('Testing', 'Basic manual testing', 'QA + automated tests', 'Comprehensive QA'),
        ('Documentation', 'Minimal', 'User guides', 'Complete docs + training'),
        ('Mobile Support', 'None', 'Responsive web', 'Native iOS/Android'),
        ('Scalability', 'Single server', 'Basic scaling', 'Enterprise-grade'),
        ('Security', 'Basic auth', 'Production security', 'Advanced security'),
        ('Support', 'None', 'Basic support', 'Full support + SLA'),
    ]
    
    for aspect, poc, mvp, full in comparisons:
        ws_compare.cell(row=row, column=1).value = aspect
        ws_compare.cell(row=row, column=1).font = Font(bold=True)
        ws_compare.cell(row=row, column=2).value = poc
        ws_compare.cell(row=row, column=3).value = mvp
        ws_compare.cell(row=row, column=4).value = full
        row += 1
    
    ws_compare.column_dimensions['A'].width = 20
    ws_compare.column_dimensions['B'].width = 30
    ws_compare.column_dimensions['C'].width = 30
    ws_compare.column_dimensions['D'].width = 30
    
    # Save workbook
    output_path = '/Users/bws/Projects/ionut/groza/docs/GROZA_Development_Breakdown.xlsx'
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    return output_path

if __name__ == '__main__':
    create_groza_breakdown()
